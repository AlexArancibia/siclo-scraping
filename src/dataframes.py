from datetime import datetime
from io import BytesIO

import pandas as pd
from dotenv import load_dotenv
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from src.drive_uploader import upload_file


def init_dataframes():
    """Initialize empty dataframes for all sheets."""
    df_disciplines = pd.DataFrame(columns=["nombre_gym", "nombre", "sede", "descripcion", "fuente"])
    df_places = pd.DataFrame(columns=["nombre_gym", "direccion_completa", "distrito", "horario_atencion", "fuente"])
    df_schedules = pd.DataFrame(columns=["nombre_gym", "sede", "nombre_clase", "instructor", "fecha", "dia_semana", "hora_inicio", "hora_fin", "fuente"])
    df_prices = pd.DataFrame(columns=["nombre_gym", "sede", "descripcion_plan", "valor", "moneda", "recurrencia", "fuente"])
    return df_disciplines, df_places, df_schedules, df_prices


def append_scraped_data(df_disciplines, df_places, df_schedules, df_prices, gym_name, scraped):
    """
    Append scraped dict data for one gym into the master dataframes.
    scraped = { "horarios": [...], "precios": [...], "disciplinas": [...] }
    """
    # Sedes
    if "ubicaciones" in scraped:
        temp = pd.DataFrame(scraped["ubicaciones"])
        temp["nombre_gym"] = gym_name
        df_places = pd.concat([df_places, temp], ignore_index=True)

    # Disciplines
    if "disciplinas" in scraped:
        temp = pd.DataFrame(scraped["disciplinas"])
        temp["nombre_gym"] = gym_name
        df_disciplines = pd.concat([df_disciplines, temp], ignore_index=True)

    # Schedules (horarios)
    if "horarios" in scraped:
        temp = pd.DataFrame(scraped["horarios"])
        temp["nombre_gym"] = gym_name
        df_schedules = pd.concat([df_schedules, temp], ignore_index=True)

    # Prices (precios)
    if "precios" in scraped:
        temp = pd.DataFrame(scraped["precios"])
        temp["nombre_gym"] = gym_name
        df_prices = pd.concat([df_prices, temp], ignore_index=True)

    return df_disciplines, df_places, df_schedules, df_prices


def create_excel_in_memory(df_disciplines, df_places, df_schedules, df_prices):
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_disciplines.to_excel(writer, sheet_name="Disciplinas", index=False)
        df_places.to_excel(writer, sheet_name="Sedes", index=False)
        df_schedules.to_excel(writer, sheet_name="Horarios", index=False)
        df_prices.to_excel(writer, sheet_name="Precios", index=False)

        autofit_excel(writer, ["Disciplinas", "Sedes", "Horarios", "Precios"])

    output.seek(0)  # rewind file pointer
    return output.getvalue()


def export_and_upload(df_disciplines, df_places, df_schedules, df_prices, folder_id):
    # 1. Filename with today's date
    filename = f"gyms-data-{datetime.now().strftime('%Y-%m-%d')}.xlsx"

    # 2. Create Excel in memory
    excel_bytes = create_excel_in_memory(df_disciplines, df_places, df_schedules, df_prices)

    # 3. Upload
    mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response = upload_file(excel_bytes, filename, mimetype, folder_id)

    return response


def autofit_excel(writer, sheet_names):
    """Auto-fit columns & wrap text in all sheets."""
    for sheet_name in sheet_names:
        ws = writer.book[sheet_name]

        # Auto-fit column widths
        for col_idx, col in enumerate(ws.columns, 1):
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # cap at 50 for safety
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # Enable wrap text for all cells
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)


if __name__ == "__main__":
    load_dotenv("../.env")
    data = {
        "disciplinas": [{"nombre": "yoga", "sede": "Miraflores", "descripcion": "hacemos yoga"}],
        "horarios": [{"sede": "1", "nombre_clase": "la clase xd", "instructor": "mariana", "fecha": "18-01-2025", "dia_semana": "lunes", "hora_inicio": "10:00", "hora_fin": "12:00",
                      "fuente": "https://gym.com/clases/yoga-1"}],
        "precios": [{"sede": "Todas", "descripcion_plan": "plan 1", "valor": 1000, "moneda": "PEN", "recurrencia": "anual", "fuente": "https://gym.com/precios/1"}]
    }
    df_disciplines, df_places, df_schedules, df_prices = init_dataframes()
    df_disciplines, df_places, df_schedules, df_prices = append_scraped_data(df_disciplines, df_places, df_schedules, df_prices, "gym1", data)
    df_disciplines, df_places, df_schedules, df_prices = append_scraped_data(df_disciplines, df_places, df_schedules, df_prices, "gym2", data)
    excel_bytes = create_excel_in_memory(df_disciplines, df_places, df_schedules, df_prices)
    exported_file = export_and_upload(df_disciplines, df_places, df_schedules, df_prices, "1M10yylyExJjh8hbtJt7iKGxVo1CkdB2H")
    print(exported_file)
